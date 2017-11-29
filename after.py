#!/usr/bin/python
import codecs
from html.parser import HTMLParser
from openpyxl import Workbook
import os


class myParser(HTMLParser):
    istable = False
    istr = False
    istd = False
    def handle_starttag(self, tag, attrs):
        def getAttr(attrlist, attrname):
            for i in attrlist:
                if attrname == i[0]:
                    return i[1]
            return None
        if tag == 'table' and getAttr(attrs, 'border') == '1':
            self.istable = True
        if tag == 'tr' and getAttr(attrs, 'bgcolor') == 'white' or getAttr(attrs, 'bgcolor') == 'beige':
            self.istr = True
        if tag == 'td':
            self.istd = True
    def handle_endtag(self, tag):
        if tag == 'table':
            self.istable = False
        if tag == 'tr':
            self.istr = False
        if tag == 'td':
            self.istd = False
    def handle_data(self, data):
        if self.istable and self.istr and self.istd and not data.isspace():
            liststr.append(data.strip())

def getIndexOfClassName(classname):
    for i in range(len(classNameList)):
        if classname == classNameList[i]:  #Got it!
            return i
    classNameList.append(classname)
    return len(classNameList) - 1

#行号， 课程名称， 分数
def fillScore(rowindex, classname, score):
    ws.cell(row = rowindex, column = getIndexOfClassName(classname) + 3).value = score

#获取到的Table内各项数据
liststr = list()
#科目列表
classNameList = list()

#新建Excel工作表
wb = Workbook()
ws = wb.active

filelist = os.listdir('result')
filelist.sort()
for i in range(len(filelist)):
    liststr.clear()
    f = codecs.open('result/' + filelist[i], 'r', 'gbk')
    par = myParser()
    par.feed(f.read())
    f.close()
    uid = int(filelist[i][:12]) #学号
    uidname = filelist[i][12:-5]    #姓名
    ws.cell(row = i + 2, column = 1).value = uid
    ws.cell(row = i + 2, column = 2).value = uidname
    for j in range(0, len(liststr), 5):
        fillScore(i + 2, liststr[j + 2], liststr[j + 3])

ws['A1'] = '学号'
ws['B1'] = '姓名'
for i in range(len(classNameList)):
    ws.cell(row = 1, column = i + 3).value = classNameList[i]
wb.save('test.xlsx')

